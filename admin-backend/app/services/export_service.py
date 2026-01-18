"""보고서 내보내기 서비스.

Excel 및 PDF 형식으로 데이터를 내보냅니다.
"""

import io
import logging
from datetime import datetime, timezone
from typing import Any

logger = logging.getLogger(__name__)


class ExportService:
    """보고서 내보내기 서비스."""

    async def export_to_excel(
        self,
        data: list[dict[str, Any]],
        columns: list[dict[str, str]],
        sheet_name: str = "Report",
        title: str | None = None,
    ) -> bytes:
        """데이터를 Excel 파일로 내보내기.

        Args:
            data: 데이터 목록
            columns: 컬럼 정의 [{"key": "field_name", "header": "Display Name"}]
            sheet_name: 시트 이름
            title: 보고서 제목 (선택)

        Returns:
            Excel 파일 바이트
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise ImportError("openpyxl 패키지가 필요합니다.")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 스타일 정의
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row_idx = 1

        # 제목 추가 (선택)
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            row_idx = 3

        # 헤더 추가
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=col["header"])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        row_idx += 1

        # 데이터 추가
        for row_data in data:
            for col_idx, col in enumerate(columns, 1):
                value = row_data.get(col["key"], "")
                # datetime 처리
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, (list, dict)):
                    value = str(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

            row_idx += 1

        # 열 너비 자동 조정
        for col_idx, col in enumerate(columns, 1):
            max_length = len(col["header"])
            for row in range(1 if not title else 3, row_idx):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # 생성 시간 추가
        ws.cell(row=row_idx + 1, column=1, value=f"생성일시: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC")

        # 파일로 저장
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    async def export_to_pdf(
        self,
        data: list[dict[str, Any]],
        columns: list[dict[str, str]],
        title: str = "Report",
        orientation: str = "portrait",
    ) -> bytes:
        """데이터를 PDF 파일로 내보내기.

        Args:
            data: 데이터 목록
            columns: 컬럼 정의 [{"key": "field_name", "header": "Display Name"}]
            title: 보고서 제목
            orientation: portrait 또는 landscape

        Returns:
            PDF 파일 바이트
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            raise ImportError("reportlab 패키지가 필요합니다.")

        buffer = io.BytesIO()

        # 페이지 설정
        page_size = landscape(A4) if orientation == "landscape" else A4
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )

        elements = []
        styles = getSampleStyleSheet()

        # 제목
        title_style = ParagraphStyle(
            name="CustomTitle",
            parent=styles["Title"],
            fontSize=16,
            spaceAfter=20,
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.5 * cm))

        # 테이블 데이터 준비
        table_data = [[col["header"] for col in columns]]

        for row_data in data:
            row = []
            for col in columns:
                value = row_data.get(col["key"], "")
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                elif isinstance(value, (list, dict)):
                    value = str(value)[:50]  # 길이 제한
                elif value is None:
                    value = ""
                row.append(str(value)[:30])  # 셀당 최대 30자
            table_data.append(row)

        # 테이블 생성
        col_widths = [
            (page_size[0] - 2 * cm) / len(columns)
        ] * len(columns)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle([
                # 헤더 스타일
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                # 데이터 스타일
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                # 테두리
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                # 행 색상 교대
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ])
        )
        elements.append(table)

        # 푸터
        elements.append(Spacer(1, 0.5 * cm))
        footer_style = ParagraphStyle(
            name="Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
        )
        elements.append(
            Paragraph(
                f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC",
                footer_style,
            )
        )

        # PDF 생성
        doc.build(elements)
        buffer.seek(0)

        return buffer.getvalue()


# =============================================================================
# 보고서 타입별 내보내기 함수
# =============================================================================


async def export_users_report(
    users: list[dict],
    format: str = "excel",
) -> bytes:
    """사용자 보고서 내보내기."""
    service = ExportService()
    columns = [
        {"key": "id", "header": "ID"},
        {"key": "nickname", "header": "닉네임"},
        {"key": "email", "header": "이메일"},
        {"key": "chips", "header": "보유 칩"},
        {"key": "is_active", "header": "활성"},
        {"key": "created_at", "header": "가입일"},
        {"key": "last_login", "header": "마지막 로그인"},
    ]

    if format == "pdf":
        return await service.export_to_pdf(
            data=users,
            columns=columns,
            title="사용자 보고서",
            orientation="landscape",
        )
    else:
        return await service.export_to_excel(
            data=users,
            columns=columns,
            sheet_name="Users",
            title="사용자 보고서",
        )


async def export_transactions_report(
    transactions: list[dict],
    format: str = "excel",
) -> bytes:
    """거래 내역 보고서 내보내기."""
    service = ExportService()
    columns = [
        {"key": "id", "header": "ID"},
        {"key": "user_id", "header": "사용자 ID"},
        {"key": "type", "header": "유형"},
        {"key": "amount", "header": "금액"},
        {"key": "status", "header": "상태"},
        {"key": "created_at", "header": "일시"},
    ]

    if format == "pdf":
        return await service.export_to_pdf(
            data=transactions,
            columns=columns,
            title="거래 내역 보고서",
            orientation="landscape",
        )
    else:
        return await service.export_to_excel(
            data=transactions,
            columns=columns,
            sheet_name="Transactions",
            title="거래 내역 보고서",
        )


async def export_audit_report(
    audit_logs: list[dict],
    format: str = "excel",
) -> bytes:
    """감사 로그 보고서 내보내기."""
    service = ExportService()
    columns = [
        {"key": "id", "header": "ID"},
        {"key": "admin_username", "header": "관리자"},
        {"key": "action", "header": "액션"},
        {"key": "target_type", "header": "대상 유형"},
        {"key": "target_id", "header": "대상 ID"},
        {"key": "ip_address", "header": "IP 주소"},
        {"key": "created_at", "header": "일시"},
    ]

    if format == "pdf":
        return await service.export_to_pdf(
            data=audit_logs,
            columns=columns,
            title="감사 로그 보고서",
            orientation="landscape",
        )
    else:
        return await service.export_to_excel(
            data=audit_logs,
            columns=columns,
            sheet_name="AuditLogs",
            title="감사 로그 보고서",
        )


async def export_revenue_report(
    revenue_data: list[dict],
    format: str = "excel",
) -> bytes:
    """수익 보고서 내보내기."""
    service = ExportService()
    columns = [
        {"key": "date", "header": "날짜"},
        {"key": "total_rake", "header": "총 레이크"},
        {"key": "total_hands", "header": "총 핸드 수"},
        {"key": "unique_players", "header": "순 플레이어 수"},
        {"key": "avg_rake_per_hand", "header": "핸드당 평균 레이크"},
    ]

    if format == "pdf":
        return await service.export_to_pdf(
            data=revenue_data,
            columns=columns,
            title="수익 보고서",
        )
    else:
        return await service.export_to_excel(
            data=revenue_data,
            columns=columns,
            sheet_name="Revenue",
            title="수익 보고서",
        )
